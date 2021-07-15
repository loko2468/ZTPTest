import os
from django.dispatch import receiver
from django.db.models.signals import post_save
from .models import File, Rate, Customer, CustomerReading
from openpyxl import load_workbook

@receiver(post_save, sender=File)
def OnFileSave(sender, **kwargs):
    inst = kwargs['instance']

    if kwargs['created'] == True:
        filepath = inst.xlsx.path
        process_data(inst.xlsx.file)
        # if os.path.isfile(filepath):
        #     os.remove(filepath)


def process_data(file):
    #file = 'D:\Ben\PycharmProjects\ZTPTest\energy\EnergyConsumptionDetail_updated.xlsx'
    try:
        wb = load_workbook(filename=file)
        rates = wb['Rate Price']
        for r in list(rates.values)[1:]:
            name = r[0]
            price_gbp_per_kwh = r[1]
            qs = Rate.objects.filter(name=name)
            if qs.count() == 0:
                Rate(name=name, price_gbp_per_kwh=price_gbp_per_kwh).save()
            else:
                qs.update(price_gbp_per_kwh=price_gbp_per_kwh)

        for s in [wb[s] for s in wb.sheetnames if 'Customer' in s]:
            name = ''
            address = ''
            meter_number = ''

            for r in list(s.rows)[0:3]:
                field = r[0].value
                value = r[1].value
                # print(field,value)
                if field == 'Customer Name':
                    name = value
                elif field == 'Customer Address':
                    address = value
                elif field == 'Meter Number':
                    meter_number = value

            qs = Customer.objects.filter(meter_number=meter_number)
            if qs.count() == 0:
                Customer(name=name, address=address, meter_number=meter_number).save()
            else:
                qs.update(name=name, address=address)

            customer = Customer.objects.get(meter_number=meter_number)

            # print(customer)
            for r in list(s.rows)[5:10]:

                user_input_rate = r[0].value

                # correct user error
                if user_input_rate == 'Weekday Day Rate':
                    user_input_rate = 'Day Rate'

                rate = Rate.objects.filter(name=user_input_rate)
                if rate.count() == 1:
                    rate = rate[0]
                    # print(rate)
                    for index, cell in enumerate(r[1:], 1):
                        # print('in')
                        if cell.value is not None:
                            number = index
                            kwh_reading = cell.value
                            # print(number, kwh_reading)

                            if CustomerReading.objects.filter(customer=customer, rate=rate, kwh_reading=kwh_reading,
                                                              number=number).count() == 0:
                                CustomerReading(customer=customer, rate=rate, kwh_reading=kwh_reading, number=number).save()
                        else:
                            continue
    except:
        pass